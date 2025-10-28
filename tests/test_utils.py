"""
Test utility functions
"""
import pytest
import openpyxl
import tempfile
import os
from utils import auto_grade_answer, parse_excel_students
from models import Question


class TestAutoGradeAnswer:
    """Test auto_grade_answer function"""

    def test_correct_mcq_answer(self, app, _db, question):
        """Test grading correct MCQ answer"""
        is_correct = auto_grade_answer(question, 'B')
        assert is_correct is True

    def test_incorrect_mcq_answer(self, app, _db, question):
        """Test grading incorrect MCQ answer"""
        is_correct = auto_grade_answer(question, 'A')
        assert is_correct is False

    def test_case_insensitive_mcq(self, app, _db, question):
        """Test that MCQ grading is case insensitive"""
        is_correct = auto_grade_answer(question, 'b')
        assert is_correct is True

    def test_correct_true_false_answer(self, app, _db, true_false_question):
        """Test grading correct True/False answer"""
        is_correct = auto_grade_answer(true_false_question, 'False')
        assert is_correct is True

    def test_incorrect_true_false_answer(self, app, _db, true_false_question):
        """Test grading incorrect True/False answer"""
        is_correct = auto_grade_answer(true_false_question, 'True')
        assert is_correct is False

    def test_case_insensitive_true_false(self, app, _db, true_false_question):
        """Test that True/False grading is case insensitive"""
        is_correct = auto_grade_answer(true_false_question, 'false')
        assert is_correct is True

    def test_whitespace_handling(self, app, _db, question):
        """Test that whitespace is stripped"""
        is_correct = auto_grade_answer(question, ' B ')
        assert is_correct is True

    def test_empty_answer(self, app, _db, question):
        """Test grading empty answer"""
        is_correct = auto_grade_answer(question, '')
        assert is_correct is False

    def test_none_answer(self, app, _db, question):
        """Test grading None answer"""
        is_correct = auto_grade_answer(question, None)
        assert is_correct is False


class TestParseExcelStudents:
    """Test parse_excel_students function"""

    def test_parse_valid_excel(self, app, _db, school_class):
        """Test parsing valid Excel file"""
        # Create a temporary Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name', 'email'])
        ws.append(['STU001', 'John', 'Doe', 'john.doe@test.com'])
        ws.append(['STU002', 'Jane', 'Smith', 'jane.smith@test.com'])

        # Save to temp file
        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            # Parse the file
            result = parse_excel_students(path, school_class.id)

            assert result['success'] is True
            assert result['created'] == 2
            assert len(result['errors']) == 0

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_excel_missing_columns(self, app, _db, school_class):
        """Test parsing Excel with missing required columns"""
        # Create Excel file missing 'email' column
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name'])
        ws.append(['STU001', 'John', 'Doe'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            result = parse_excel_students(path, school_class.id)

            assert result['success'] is False
            assert 'missing columns' in result['message'].lower() or 'required' in result['message'].lower()

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_excel_duplicate_student_id(self, app, _db, school_class):
        """Test parsing Excel with duplicate student IDs"""
        # Create Excel with duplicate student_id
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name', 'email'])
        ws.append(['STU001', 'John', 'Doe', 'john@test.com'])
        ws.append(['STU001', 'Jane', 'Smith', 'jane@test.com'])  # Duplicate ID

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            result = parse_excel_students(path, school_class.id)

            # Should have errors for duplicate
            assert len(result['errors']) > 0

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_excel_duplicate_email(self, app, _db, school_class):
        """Test parsing Excel with duplicate emails"""
        # Create Excel with duplicate email
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name', 'email'])
        ws.append(['STU001', 'John', 'Doe', 'same@test.com'])
        ws.append(['STU002', 'Jane', 'Smith', 'same@test.com'])  # Duplicate email

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            result = parse_excel_students(path, school_class.id)

            # Should have errors for duplicate
            assert len(result['errors']) > 0

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_excel_empty_file(self, app, _db, school_class):
        """Test parsing empty Excel file"""
        # Create empty Excel file (only headers)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name', 'email'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            result = parse_excel_students(path, school_class.id)

            assert result['success'] is True
            assert result['created'] == 0

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_excel_invalid_email(self, app, _db, school_class):
        """Test parsing Excel with invalid email"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name', 'email'])
        ws.append(['STU001', 'John', 'Doe', 'notanemail'])  # Invalid email

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            result = parse_excel_students(path, school_class.id)

            # Should have errors for invalid email
            assert len(result['errors']) > 0 or result['created'] == 0

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_excel_missing_values(self, app, _db, school_class):
        """Test parsing Excel with missing values in rows"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name', 'email'])
        ws.append(['STU001', '', 'Doe', 'john@test.com'])  # Missing first_name
        ws.append(['STU002', 'Jane', 'Smith', 'jane@test.com'])  # Valid

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            result = parse_excel_students(path, school_class.id)

            # Should skip row with missing required field
            assert result['created'] <= 2
            if result['created'] < 2:
                assert len(result['errors']) > 0

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_excel_creates_users(self, app, _db, school_class):
        """Test that parsing Excel creates both User and Student records"""
        from models import User, Student

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name', 'email'])
        ws.append(['STU999', 'Test', 'Student', 'test.student@test.com'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            result = parse_excel_students(path, school_class.id)

            assert result['success'] is True
            assert result['created'] == 1

            # Verify User created
            user = User.query.filter_by(email='test.student@test.com').first()
            assert user is not None
            assert user.role == 'student'
            assert user.is_active is True

            # Verify Student created
            student = Student.query.filter_by(student_id='STU999').first()
            assert student is not None
            assert student.first_name == 'Test'
            assert student.last_name == 'Student'
            assert student.class_id == school_class.id

        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_parse_nonexistent_file(self, app, _db, school_class):
        """Test parsing non-existent file"""
        result = parse_excel_students('/nonexistent/file.xlsx', school_class.id)

        assert result['success'] is False
        assert 'error' in result['message'].lower() or 'not found' in result['message'].lower()

    def test_parse_invalid_class_id(self, app, _db):
        """Test parsing with invalid class ID"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['student_id', 'first_name', 'last_name', 'email'])
        ws.append(['STU001', 'John', 'Doe', 'john@test.com'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        try:
            wb.save(path)
            os.close(fd)

            result = parse_excel_students(path, 99999)  # Non-existent class ID

            assert result['success'] is False or result['created'] == 0

        finally:
            if os.path.exists(path):
                os.unlink(path)


class TestParseQuestionsFromWord:
    """Test parse_questions_from_word function (if implemented)"""

    def test_parse_word_function_exists(self):
        """Test that parse_questions_from_word function exists"""
        from utils import parse_questions_from_word
        assert callable(parse_questions_from_word)


class TestParseQuestionsFromPDF:
    """Test parse_questions_from_pdf function (if implemented)"""

    def test_parse_pdf_function_exists(self):
        """Test that parse_questions_from_pdf function exists"""
        from utils import parse_questions_from_pdf
        assert callable(parse_questions_from_pdf)
